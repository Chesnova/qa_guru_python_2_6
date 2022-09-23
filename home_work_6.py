import zipfile, os
from PyPDF2 import PdfReader
from openpyxl import load_workbook


def test_create_file_zip():
    file_zip = zipfile.ZipFile('./resources/archiv.zip', 'w')
    for folder, subfolders, files in os.walk('./resources/'):
        for file in files:
            if file.endswith('.pdf') or file.endswith('.csv') or file.endswith('.xlsx'):
                file_zip.write(os.path.join(folder, file), os.path.relpath(os.path.join(folder, file), './resources/'), compress_type=zipfile.ZIP_DEFLATED)
    file_zip.close()


def test_read_pdf_file():
    with zipfile.ZipFile('./resources/archiv.zip') as file_zip:
        text_pdf = file_zip.open('docs-pytest-org-en-latest.pdf')
        pdf_reader = PdfReader(text_pdf)
        numbers_of_pages = len(pdf_reader.pages)
        print(numbers_of_pages)
        assert numbers_of_pages == 412
        page = pdf_reader.pages[0]
        text = page.extract_text()
        print(text)
        assert 'Jul 11, 2022' in text


def test_read_csv_file():
    with zipfile.ZipFile('./resources/archiv.zip') as file_zip:
        csv_file = file_zip.read('SampleCSVFile_11kb.csv').decode('utf-8')
        assert 'Eldon Base' in csv_file
        row_count = sum(1 for row in csv_file)
        assert row_count == 10996


def test_read_xlsx_file():
    with zipfile.ZipFile('./resources/archiv.zip') as file_zip:
        work_book = file_zip.open('file_example_XLSX_50.xlsx')
        workbook = load_workbook(work_book)
        sheet = workbook.active
        text = sheet.cell(row=3, column=2).value
        assert text in 'Mara'


